import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class Library:
    def __init__(self, root):
        self.root = root
        self.root.title("Library Management System")
        
        # Frame Setup
        self.frame1 = tk.Frame(self.root)
        self.frame1.pack(side=tk.TOP, fill=tk.X)
        
        self.frame2 = tk.Frame(self.root)
        self.frame2.pack(side=tk.BOTTOM, fill=tk.BOTH, expand=True)
        
        # Input Fields
        self.book_label = tk.Label(self.frame1, text="Book Name:", font=("Arial", 14))
        self.book_label.grid(row=0, column=0, pady=5, padx=5)
        self.book_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.book_input.grid(row=0, column=1, pady=5, padx=5)
        
        self.rollno_label = tk.Label(self.frame1, text="Roll No:", font=("Arial", 14))
        self.rollno_label.grid(row=0, column=2, pady=5, padx=5)
        self.rollno_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.rollno_input.grid(row=0, column=3, pady=5, padx=5)

        # Additional Fields
        self.student_name_label = tk.Label(self.frame1, text="Student Name:", font=("Arial", 14))
        self.student_name_label.grid(row=1, column=0, pady=5, padx=5)
        self.student_name_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.student_name_input.grid(row=1, column=1, pady=5, padx=5)

        self.course_label = tk.Label(self.frame1, text="Course Name:", font=("Arial", 14))
        self.course_label.grid(row=1, column=2, pady=5, padx=5)
        self.course_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.course_input.grid(row=1, column=3, pady=5, padx=5)

        self.branch_label = tk.Label(self.frame1, text="Branch Name:", font=("Arial", 14))
        self.branch_label.grid(row=2, column=0, pady=5, padx=5)
        self.branch_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.branch_input.grid(row=2, column=1, pady=5, padx=5)

        self.mobile_label = tk.Label(self.frame1, text="Mobile Number:", font=("Arial", 14))
        self.mobile_label.grid(row=2, column=2, pady=5, padx=5)
        self.mobile_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.mobile_input.grid(row=2, column=3, pady=5, padx=5)

        self.date_label = tk.Label(self.frame1, text="Date of Issue (YYYY-MM-DD):", font=("Arial", 14))
        self.date_label.grid(row=3, column=0, pady=5, padx=5)
        self.date_input = tk.Entry(self.frame1, font=("Arial", 14))
        self.date_input.grid(row=3, column=1, pady=5, padx=5)
        
        # Allocate Book Button
        self.search_button = tk.Button(self.frame1, text="Allocate Book", font=("Arial", 14), command=self.allocate_book)
        self.search_button.grid(row=4, column=0, columnspan=4, pady=15)
        
        # Treeview for displaying allocated books
        self.tree = ttk.Treeview(self.frame2, columns=("Book Name", "Roll No", "Student Name", "Course Name", "Branch Name", "Mobile Number", "Date of Issue"), show="headings", height=15)
        
        # Define column headings
        self.tree.heading("Book Name", text="Book Name")
        self.tree.heading("Roll No", text="Roll No")
        self.tree.heading("Student Name", text="Student Name")
        self.tree.heading("Course Name", text="Course Name")
        self.tree.heading("Branch Name", text="Branch Name")
        self.tree.heading("Mobile Number", text="Mobile Number")
        self.tree.heading("Date of Issue", text="Date of Issue")
        
        # Define column widths
        self.tree.column("Book Name", width=100)
        self.tree.column("Roll No", width=100)
        self.tree.column("Student Name", width=150)
        self.tree.column("Course Name", width=100)
        self.tree.column("Branch Name", width=100)
        self.tree.column("Mobile Number", width=120)
        self.tree.column("Date of Issue", width=120)
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Sample Books List (20 B.Tech CSE Books)
        self.books = [
            "Data Structures and Algorithms in C", 
            "Operating System Concepts", 
            "Database Management Systems", 
            "Computer Networks", 
            "Discrete Mathematics", 
            "Programming in C", 
            "Software Engineering", 
            "Artificial Intelligence", 
            "Data Structures and Algorithms", 
            "Compiler Design", 
            "Computer Architecture", 
            "Digital Logic Design", 
            "Java Programming", 
            "Web Technologies", 
            "Computer Graphics", 
            "Microprocessor and Microcontroller", 
            "Design and Analysis of Algorithms", 
            "Machine Learning", 
            "Computer Organization", 
            "Cyber Security"
        ]
        
        # Sample Student Data (Roll numbers from 100 to 200)
        self.students = [str(roll) for roll in range(100, 201)]

        # Excel file setup
        self.filename = "LibraryData.xlsx"
        if not os.path.exists(self.filename):
            self.create_excel_file()

    def create_excel_file(self):
        # Create a new Excel file and add headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Allocations"
        headers = ["Book Name", "Roll No", "Student Name", "Course Name", "Branch Name", "Mobile Number", "Date of Issue"]
        sheet.append(headers)
        workbook.save(self.filename)

    def allocate_book(self):
        # Get input values
        book_name = self.book_input.get().strip()
        roll_no = self.rollno_input.get().strip()
        student_name = self.student_name_input.get().strip()
        course_name = self.course_input.get().strip()
        branch_name = self.branch_input.get().strip()
        mobile_number = self.mobile_input.get().strip()
        date_of_issue = self.date_input.get().strip()

        # Check if the book and roll number are valid
        book_found = book_name in self.books
        rollno_found = roll_no in self.students
        
        if book_found and rollno_found:
            # Insert new record in Treeview
            self.tree.insert("", tk.END, values=(book_name, roll_no, student_name, course_name, branch_name, mobile_number, date_of_issue))
            self.save_to_excel(book_name, roll_no, student_name, course_name, branch_name, mobile_number, date_of_issue)
        else:
            if not book_found:
                messagebox.showerror("Error", "Book Not Found")
            if not rollno_found:
                messagebox.showerror("Error", "Invalid Roll No")
        
        # Clear input fields after allocation
        self.clear_inputs()

    def save_to_excel(self, book_name, roll_no, student_name, course_name, branch_name, mobile_number, date_of_issue):
        # Load the workbook and add the new record to the sheet
        workbook = load_workbook(self.filename)
        sheet = workbook["Allocations"]
        sheet.append([book_name, roll_no, student_name, course_name, branch_name, mobile_number, date_of_issue])
        workbook.save(self.filename)

    def clear_inputs(self):
        # Clear all input fields
        self.book_input.delete(0, tk.END)
        self.rollno_input.delete(0, tk.END)
        self.student_name_input.delete(0, tk.END)
        self.course_input.delete(0, tk.END)
        self.branch_input.delete(0, tk.END)
        self.mobile_input.delete(0, tk.END)
        self.date_input.delete(0, tk.END)

# Run Application
root = tk.Tk()
app = Library(root)
root.mainloop()